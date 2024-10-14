from flask import Flask, render_template, request, send_file
from openpyxl import load_workbook
import io
import os
import logging

app = Flask(__name__)

# Para registrar errores
logging.basicConfig(filename='error.log', level=logging.ERROR)

# Ruta inicial que carga el formulario
@app.route('/')
def index():
    return render_template('index.html')

# Procesa el archivo y muestra los checkboxes
@app.route('/generate_checkboxes', methods=['POST'])
def generate_checkboxes():
    try:
        # Verificar si hay un archivo en la petición
        if 'file' not in request.files:
            return "No se ha enviado ningún archivo", 400

        file = request.files['file']
        if file.filename == '':
            return "Archivo no seleccionado", 400

        # Guardar temporalmente el archivo
        file_path = f'/tmp/{file.filename}'
        file.save(file_path)

        # Cargar el archivo Excel y obtener los nombres
        wb = load_workbook(file_path)
        ws = wb.active

        # Lista de personas a mostrar en los checkboxes
        people = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] and row[1]:  # Asegurarse de que Name y Surname no sean nulos
                people.append({
                    'index': i,        # Índice de la fila
                    'name': row[0],    # Columna A ("Name")
                    'surname': row[1]  # Columna B ("Surname")
                })

        # Renderizar la página con los checkboxes
        return render_template('index.html', people=people, file_path=file_path)

    except Exception as e:
        logging.error(f'Error: {e}')
        return "Se produjo un error interno", 500

# Procesa la selección y exporta los datos seleccionados
@app.route('/export_selected', methods=['POST'])
def export_selected():
    try:
        selected_people = request.form.getlist('selected_people')
        file_path = request.form['file_path']

        if not selected_people:
            return "No seleccionaste ninguna persona", 400

        # Cargar la plantilla original
        plantilla_path = 'PlantillaSTEP4.xlsx'
        wb_plantilla = load_workbook(plantilla_path)
        ws_plantilla = wb_plantilla.active

        # Cargar el archivo original
        wb = load_workbook(file_path)
        ws = wb.active

        # Transferir los datos seleccionados a la plantilla
        for index in selected_people:
            row_index = int(index)
            row = list(ws.iter_rows(min_row=row_index, max_row=row_index, values_only=True))[0]

            # Asignar datos a la plantilla, comenzando desde la fila 7
            plantilla_row = row_index + 5  # Ajuste para comenzar desde la fila 7 en la plantilla
            ws_plantilla[f'C{plantilla_row}'] = row[0]  # Columna "Name"
            ws_plantilla[f'D{plantilla_row}'] = row[1]  # Columna "Surname"
            ws_plantilla[f'E{plantilla_row}'] = row[14] if len(row) > 14 else ""  # Columna "Primary email"
            ws_plantilla[f'F{plantilla_row}'] = row[18] if len(row) > 18 else ""  # Columna "Primary phone"
            ws_plantilla[f'G{plantilla_row}'] = 'D_PCC'  # Ejemplo de asignación de "Workgroup"
            ws_plantilla[f'H{plantilla_row}'] = 'Team_D_CCH_PCC_1'  # Ejemplo de asignación de "Team"
            ws_plantilla[f'L{plantilla_row}'] = row[4] if len(row) > 4 else ""  # "Is PCC"
            ws_plantilla[f'Q{plantilla_row}'] = row[15] if len(row) > 15 else ""  # "CTI User"
            ws_plantilla[f'R{plantilla_row}'] = row[15] if len(row) > 15 else ""  # "TTG UserID 1"
            ws_plantilla[f'V{plantilla_row}'] = 'Agent'  # "Campaign Level"

        # Guardar el archivo generado en memoria
        output_stream = io.BytesIO()
        wb_plantilla.save(output_stream)
        output_stream.seek(0)

        # Eliminar el archivo temporal
        os.remove(file_path)

        # Enviar el archivo generado al cliente
        return send_file(output_stream, as_attachment=True, download_name="Plantilla_generada.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        logging.error(f'Error: {e}')
        return "Se produjo un error interno", 500

if __name__ == '__main__':
    app.run(debug=True)
