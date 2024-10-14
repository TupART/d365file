from flask import Flask, request, render_template, send_file, redirect, url_for
import openpyxl
from procesar_excel import procesar_datos
import os

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/generate_checkboxes', methods=['POST'])
def generate_checkboxes():
    file = request.files['file']
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    
    # Empieza a leer desde la fila 3
    data = []
    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=2, values_only=True):
        name, surname = row
        if name and surname:
            data.append({'name': name, 'surname': surname})
    
    if not data:
        return 'No se encontraron personas en el archivo', 400
    
    return render_template('checkboxes.html', people=data)

@app.route('/process_selected', methods=['POST'])
def process_selected():
    selected = request.form.getlist('selected_people')
    
    file = request.files['file']
    if not selected:
        return 'No seleccionaste ninguna persona.', 400
    
    # Llama a la función para procesar y generar la plantilla con las personas seleccionadas
    output_path = procesar_datos(file, selected)
    
    return send_file(output_path, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
